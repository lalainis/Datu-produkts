import argparse
import json
import math
import re
from collections import Counter, defaultdict
from datetime import UTC, datetime
from pathlib import Path

import openpyxl


BASE_DIR = Path(__file__).resolve().parent.parent
DEFAULT_CONSUMPTION_FILE = BASE_DIR / "ofisu komplekss.xlsx"
PRICE_FILE = BASE_DIR / "NP_Cenas_LV.xlsx"
OUTPUT_FILE = BASE_DIR / "data" / "app-data.json"

INTERVAL_RE = re.compile(r"^\d{2}:\d{2} - \d{2}:\d{2}$")


def round_up(value, step=5):
    return int(math.ceil(value / step) * step)


def percentile(values, ratio):
    if not values:
        return 0
    if len(values) == 1:
        return values[0]
    position = (len(values) - 1) * ratio
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return values[lower]
    fraction = position - lower
    return values[lower] + (values[upper] - values[lower]) * fraction


def iso_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value)[:10]


def hour_label(interval_text):
    start = str(interval_text).split("-")[0].strip()
    return f"{start}:00"


def price_hour_label(interval_text):
    start = str(interval_text).split(" - ")[0].strip()
    return f"{start[:2]}:00"


def detect_consumption_layout(header):
    first_cell = str(header[0] or "").strip().lower()
    if first_cell.startswith("obj"):
        return "portfolio"
    if first_cell.startswith("dat"):
        return "simple"
    raise ValueError(f"Unsupported consumption sheet layout: {header}")


def parse_consumption_row(row, layout, fallback_name):
    if layout == "portfolio":
        if not row[1]:
            return None
        allowed_load = row[5] if len(row) > 5 and row[5] is not None and len(row) > 8 else row[4]
        market_price_index = 7 if len(row) > 8 else 6
        return {
            "object_name": str(row[0]).split(" (")[0] if row[0] else fallback_name,
            "record_date": iso_date(row[1]),
            "interval": str(row[2]),
            "consumption": float(row[3] or 0),
            "grid_export": 0.0,
            "market_price": float(row[market_price_index] or 0),
            "allowed_load": float(allowed_load or 0),
        }

    if not row[0]:
        return None
    return {
        "object_name": fallback_name,
        "record_date": iso_date(row[0]),
        "interval": str(row[1]),
        "consumption": float(row[2] or 0),
        "grid_export": float(row[3] or 0),
        "market_price": float(row[4] or 0),
        "allowed_load": None,
    }


def load_consumption_data(consumption_file):
    workbook = openpyxl.load_workbook(consumption_file, data_only=True, read_only=True)
    objects = []

    for worksheet in workbook.worksheets:
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows)
        layout = detect_consumption_layout(header)
        fallback_name = worksheet.title if len(workbook.worksheets) > 1 else consumption_file.stem
        object_name = None
        dates = []
        consumptions = []
        allowed_values = []
        daily_totals = defaultdict(lambda: {"consumption": 0.0, "cost": 0.0})
        hourly_profile = defaultdict(list)
        export_profile = defaultdict(list)
        anomalies = []
        grid_exports = []

        for row in rows:
            parsed_row = parse_consumption_row(row, layout, fallback_name)
            if parsed_row is None:
                continue

            if object_name is None:
                object_name = parsed_row["object_name"]

            dates.append(parsed_row["record_date"])
            consumptions.append(parsed_row["consumption"])
            grid_exports.append(parsed_row["grid_export"])
            if parsed_row["allowed_load"] is not None:
                allowed_values.append(parsed_row["allowed_load"])

            day_data = daily_totals[parsed_row["record_date"]]
            day_data["consumption"] += parsed_row["consumption"]
            day_data["cost"] += (parsed_row["consumption"] * parsed_row["market_price"]) / 1000
            hourly_profile[hour_label(parsed_row["interval"])].append(parsed_row["consumption"])
            if parsed_row["grid_export"] > 0:
                export_profile[hour_label(parsed_row["interval"])].append(parsed_row["grid_export"])

        sorted_consumptions = sorted(consumptions)
        mean_consumption = sum(consumptions) / len(consumptions)
        variance = sum((value - mean_consumption) ** 2 for value in consumptions) / len(consumptions)
        std_deviation = math.sqrt(variance)
        anomaly_threshold = mean_consumption + 2 * std_deviation
        if allowed_values:
            current_allowed_load = Counter(allowed_values).most_common(1)[0][0]
        else:
            current_allowed_load = round_up(max(percentile(sorted_consumptions, 0.98) * 1.05, max(consumptions) * 1.1))
        recommended_allowed_load = round_up(max(max(consumptions) * 1.05, percentile(sorted_consumptions, 0.99) * 1.1))
        recommended_allowed_load = max(recommended_allowed_load, current_allowed_load)

        rows = worksheet.iter_rows(min_row=2, values_only=True)
        for row in rows:
            parsed_row = parse_consumption_row(row, layout, fallback_name)
            if parsed_row is None:
                continue
            reasons = []
            if parsed_row["consumption"] > current_allowed_load:
                reasons.append("Pārsniedz atļauto slodzi")
            elif parsed_row["consumption"] >= current_allowed_load * 0.9:
                reasons.append("Tuvu atļautajai slodzei")
            if parsed_row["consumption"] >= anomaly_threshold:
                reasons.append("Anomāli augsts patēriņš")
            if reasons:
                anomalies.append(
                    {
                        "date": parsed_row["record_date"],
                        "hour": hour_label(parsed_row["interval"]),
                        "consumption": round(parsed_row["consumption"], 2),
                        "reason": ", ".join(reasons),
                    }
                )

        hourly_profile_items = [
            {
                "hour": hour,
                "consumption": round(sum(values) / len(values), 2),
            }
            for hour, values in sorted(hourly_profile.items())
        ]
        export_profile_items = [
            {
                "hour": hour,
                "export": round(sum(values) / len(values), 2),
            }
            for hour, values in sorted(export_profile.items())
        ]
        daily_totals_items = [
            {
                "date": date,
                "consumption": round(values["consumption"], 2),
                "cost": round(values["cost"], 2),
            }
            for date, values in sorted(daily_totals.items())
        ]
        total_consumption = round(sum(consumptions), 2)
        total_cost = round(sum(item["cost"] for item in daily_totals_items), 2)
        total_export = round(sum(grid_exports), 2)
        peak_export = round(max(grid_exports), 2) if grid_exports else 0.0
        export_hours = sum(1 for value in grid_exports if value > 0)
        solar_detected = total_export > 0 or "ses" in consumption_file.stem.lower()

        objects.append(
            {
                "id": worksheet.title,
                "name": object_name or worksheet.title,
                "hourlyProfile": hourly_profile_items,
                "exportHourlyProfile": export_profile_items,
                "last30Days": daily_totals_items[-30:],
                "anomalies": sorted(anomalies, key=lambda item: item["consumption"], reverse=True)[:20],
                "solar": {
                    "detected": solar_detected,
                    "totalExport": total_export,
                    "peakExport": peak_export,
                    "exportHours": export_hours,
                    "averageDailyExport": round(total_export / len(daily_totals_items), 2) if daily_totals_items else 0.0,
                    "topExportHours": sorted(export_profile_items, key=lambda item: item["export"], reverse=True)[:5],
                },
                "summary": {
                    "periodStart": min(dates),
                    "periodEnd": max(dates),
                    "totalConsumption": total_consumption,
                    "totalCost": total_cost,
                    "averageHourlyConsumption": round(mean_consumption, 2),
                    "averageDailyConsumption": round(total_consumption / len(daily_totals_items), 2),
                    "peakHourlyConsumption": round(max(consumptions), 2),
                    "currentAllowedLoadKw": round(current_allowed_load, 2),
                    "recommendedAllowedLoadKw": round(recommended_allowed_load, 2),
                    "anomalyCount": len(anomalies),
                    "hoursAboveRecommendedLoad": sum(1 for value in consumptions if value > recommended_allowed_load),
                    "topPeakHours": sorted(hourly_profile_items, key=lambda item: item["consumption"], reverse=True)[:5],
                },
            }
        )

    return objects


def load_price_data():
    workbook = openpyxl.load_workbook(PRICE_FILE, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    date_row = rows[5]
    date_columns = []

    for index in range(2, len(date_row)):
        if isinstance(date_row[index], datetime):
            date_columns.append((index, date_row[index].date().isoformat()))

    prices_by_date = {date: [] for _, date in date_columns}

    for row in rows[7:]:
        interval = row[1]
        if not INTERVAL_RE.match(str(interval or "")):
            continue
        for index, date in date_columns:
            price = row[index]
            if price is not None:
                prices_by_date[date].append({"interval": interval, "price": float(price)})

    daily_averages = []
    hourly_by_date = {}
    for date, quarter_hours in prices_by_date.items():
        grouped = defaultdict(list)
        for item in quarter_hours:
            hour = price_hour_label(item["interval"])
            grouped[hour].append(item["price"])
        hourly_items = [
            {"hour": hour, "price": round(sum(values) / len(values), 2)}
            for hour, values in sorted(grouped.items())
        ]
        hourly_by_date[date] = hourly_items
        if hourly_items:
            price_values = [item["price"] for item in hourly_items]
            daily_averages.append(
                {
                    "date": date,
                    "averagePrice": round(sum(price_values) / len(price_values), 2),
                    "minPrice": round(min(price_values), 2),
                    "maxPrice": round(max(price_values), 2),
                }
            )

    latest_date = max(hourly_by_date)
    latest_hourly = hourly_by_date[latest_date]
    cheapest_hours = sorted(latest_hourly, key=lambda item: item["price"])[:3]
    expensive_hours = sorted(latest_hourly, key=lambda item: item["price"], reverse=True)[:3]

    return {
        "marketPricePeriod": {
            "startDate": min(prices_by_date),
            "endDate": max(prices_by_date),
        },
        "tomorrowPrices": {
            "date": latest_date,
            "hourly": latest_hourly,
            "averagePrice": round(sum(item["price"] for item in latest_hourly) / len(latest_hourly), 2),
            "cheapestHours": sorted(cheapest_hours, key=lambda item: item["hour"]),
            "expensiveHours": sorted(expensive_hours, key=lambda item: item["hour"]),
        },
        "priceHistory": daily_averages,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=str(DEFAULT_CONSUMPTION_FILE))
    args = parser.parse_args()

    consumption_file = Path(args.source)
    OUTPUT_FILE.parent.mkdir(exist_ok=True)
    price_data = load_price_data()
    result = {
        "generatedAt": datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "sourceFile": consumption_file.name,
        "sourceHasSolar": "ses" in consumption_file.stem.lower(),
        "marketPricePeriod": price_data["marketPricePeriod"],
        "tomorrowPrices": price_data["tomorrowPrices"],
        "priceHistory": price_data["priceHistory"],
        "objects": load_consumption_data(consumption_file),
    }
    result["sourceHasSolar"] = result["sourceHasSolar"] or any(item["solar"]["detected"] for item in result["objects"])
    OUTPUT_FILE.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Izveidots {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
